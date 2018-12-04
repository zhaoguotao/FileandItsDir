#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  fileParserOpenpyxl.py
#
#  Copyright 2018 赵国涛 <guotao.zhao@vivo.com>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#



import os
import time
import sys
#~ import xlwt
from openpyxl import Workbook, load_workbook

#~ print(os.path.abspath('.'))
file_dir = "." #定义文件目录为当前py文件所在目录

class FileCheck():

    def __init__(self):
        self.file_dir = file_dir

    def get_filesize(self,filename):
        """
        Get file size（M: MB）
        """
        file_byte = os.path.getsize(filename)
        return self.sizeConvert(file_byte)


    def get_file_Date(self,filename):
        """
        Get file date
        """
        filemt=time.localtime(os.stat(filename).st_mtime)
        fmt=time.strftime("%Y.%m.%d_%H:%M:%S",filemt)
        return fmt

    def sizeConvert(self,size):# 单位换算
        K, M, G = 1024, 1024**2, 1024**3
        if size >= G:
            return str(round(size/G, 2))+'GB'
        elif size >= M:
            return str(round(size/M,2))+'MB'
        elif size >= K:
            return str(round(size/K,2))+'KB'
        else:
            return str(size)+'B'

    def get_all_file(self,filepath):
        """
        Get all files
        """
        for root, dirs, files in os.walk(file_dir):
            return files


print(">>>>>>>>>>>>>>>>>start...")
fc = FileCheck()
datas = [['文件名称', '文件大小', '修改日期',"文件路径"]]#二维数组

def gci(filepath):
    files = os.listdir(filepath)
    for fi in files:
        cell = []
        print(fi)
        fi_d = os.path.join(filepath,fi)
        if os.path.isdir(fi_d):
            gci(fi_d)
        else:
            file_path     = fi_d
            file_size     = fc.get_filesize(file_path)
            file_date     = fc.get_file_Date(file_path)
            cell.append(fi)#file name
            cell.append(file_size)
            cell.append(file_date)
            cell.append(file_path)
            datas.append(cell)


print(os.path.abspath('.'))
gci(file_dir)


wb = Workbook() #Create xlsx file
sheet = wb.active
#sheet2=wb.create_sheet('test')#sheet的名称为test

for row in range(0, len(datas)):
    sheet.append(datas[row])
wb.save("FileParserResults.xlsx")
print("============Finished================")
